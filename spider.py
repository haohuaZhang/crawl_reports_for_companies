from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.firefox import GeckoDriverManager
from bs4 import BeautifulSoup
import pandas as pd
import re
import os
import random
import time
from io import StringIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
import requests
from requests.exceptions import ConnectionError

# 调整所有 sheet 的列宽
def auto_adjust_column_width(excel_file_path):
    # 打开 Excel 文件
    workbook = load_workbook(excel_file_path)
    
    # 遍历每个 sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 遍历每一列，调整列宽
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    # 获取每个单元格的值的长度
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # 设置列宽为最大长度 + 2（留一些额外空间）
            adjusted_width = max_length + 20
            sheet.column_dimensions[column].width = adjusted_width

    # 保存调整后的文件
    workbook.save(excel_file_path)

  
# 构建公司公告页面URL
def build_announcement_urls(stock_id, reportTypes):
    urls = [];
    for currentType in reportTypes:
        url = f'https://vip.stock.finance.sina.com.cn/corp/go.php/vCB_Bulletin/stockid/{stock_id}/page_type/{currentType}.phtml'
        urls.append(url)

    return urls

# 配置 Firefox 浏览器选项
options = Options()
options.set_preference('permissions.default.image', 2)  # 禁用图片
options.add_argument('--headless')  # 在 Colab 环境中无界面运行
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')


# 指定 geckodriver 的路径，使用 Service 来设置
service = Service('/usr/local/bin/geckodriver')
# service = Service(GeckoDriverManager().install())
# service = Service('D:\\work\\GitHub\\files\\geckodriver.exe')
# 动态获取，不需要像上面一样判断系统是mac还是window
# service = Service(GeckoDriverManager().install(), port=0)


# 创建 WebDriver 实例
driver = webdriver.Firefox(service=service, options=options)
# 使用 webdriver_manager 动态下载和配置 GeckoDriver
# driver = webdriver.Firefox(service=Service(service), options=options)

def get_reports_urls(soup, years):
    # 存放符合条件的链接
    links = []

    # 查找 class="datelist" 的 div
    datelist_div = soup.find('div', class_='datelist')
    
    if datelist_div:
        # 查找 div 内所有 a 标签
        a_tags = datelist_div.find_all('a', href=True)

        # 遍历所有 a 标签
        for a_tag in a_tags:
            # 获取链接文本
            link_text = a_tag.text.strip()
            
            # 检查文本是否包含“半年度报告”或“年度报告”，且年份在 years 数组中
            for year in years:
                if ('半年度报告' in link_text or '年度报告' in link_text or '三季度报告' in link_text or '一季度报告' in link_text ) and str(year) in link_text:
                    # 满足条件的链接加入 links 列表
                    links.append(a_tag)
    
    return links

# 获取公司最新的年度报告和半年度报告链接
def get_latest_reports_urls(stock_id, years, reportTypes):
    urls = build_announcement_urls(stock_id, reportTypes)
    report_urls = {}  # 初始化字典，避免在异常情况下未定义
    for url in urls:
        driver.get(url)

        try:
            # 等待页面加载完成
            WebDriverWait(driver, 5).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.datelist'))
            )
            
            # 获取页面源码
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')

            # 查找符合条件的报告链接
            links = get_reports_urls(soup, years)

            # 检查链接是否为 BeautifulSoup 对象
            for link in links:
                if isinstance(link, str):
                    # 如果 link 是字符串，则解析为 <a> 标签
                    link = BeautifulSoup(link, 'html.parser').a
                text = link.text.strip()  # 获取 <a> 标签的文本
                href = 'https://vip.stock.finance.sina.com.cn/' + link['href']       # 获取 href 属性
                report_urls[text] = href  # 将文本作为 key，href 作为 value


        except Exception as e:
            print(f"获取报告链接时出现错误: {e}")
            # 返回一个空字典，而不是 None
            report_urls = {}

    return report_urls



# 目标目录
directory = './dist'

# 初始化存放 reports 文件名的列表
reports_files = []

# 正则表达式匹配以 'reports' 开头，后面可选数字，并以 '.xlsx' 结尾的文件
pattern = r'^reports(\d*)\.xlsx$'

# 遍历 dist 目录下的文件
for filename in os.listdir(directory):
    match = re.match(pattern, filename)
    if match:
        # 提取文件名中的数字部分
        num_str = match.group(1)
        # 如果没有数字部分，默认为 0
        number = int(num_str) if num_str else 0
        reports_files.append(number)

# 找到当前最大数字
if reports_files:
    max_number = max(reports_files)
else:
    max_number = 0

# 下一个文件名
new_filename = f'reports{max_number + 1}.xlsx'

# 完整的导出路径
output_file = os.path.join(directory, new_filename)


# 继续执行保存 Excel 的操作，例如：
# df.to_excel(output_file)


# 主函数，爬取多个公司的报告并提取研发费用信息
# 示例：如何调用函数将不同的报告保存到同一个 Excel 文件
def crawl_reports_for_companies(companies, years, target_tables, reportTypes = ['zqbg', 'ndbg', 'sjdbg']):
    results = []
    
    # 创建 ExcelWriter 对象，用于写入多个 sheet
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for company in companies:
            stock_id = get_stock_code_by_company_name(company)

            if not stock_id:
                print(f"无法找到公司 {company} 的股票代码")
                continue

            report_urls = get_latest_reports_urls(stock_id, years, reportTypes)
            print(f"stock_id {stock_id}  report_urls {report_urls}")

            if not report_urls:
                print(f"未找到 {company} 的报告链接")
                continue

            for report_type, report_url in report_urls.items():
                if report_url:
                    print(f"正在爬取 {company} 的 {report_type}...")
                    sheet_name = f"{report_type}"
                    # 调用带重试机制的爬取函数，将数据写入 Excel 不同 sheet
                    report = get_report_content_selenium(sheet_name, report_url, writer, target_tables)
                    
                    if report:
                        print(f"{company} 的 {report_type} 报告爬取成功")
                        results.append(report)
                    else:
                        print(f"无法爬取 {company} 的 {report_type} 报告")
                    
                    # 在每次请求之间添加随机延迟
                    time.sleep(random.uniform(3, 7))  # 休息 3 到 7 秒以避免被检测为爬虫
                else:
                    print(f"未找到 {company} 的 {report_type} 链接")
    driver.quit()  # 确保浏览器关闭
    # 调整列宽
    auto_adjust_column_width(output_file)
    return results

# 使用Selenium爬取报告页面内容，带有重试机制
def get_report_content_selenium(sheet_name, report_url, writer, target_tables=['合并利润表'], retries=3):
    """
    爬取报告页面内容，并将表格数据写入到 Excel 的同一个 sheet。
    
    :param report_url: 报告的 URL
    :param writer: pd.ExcelWriter 对象，用于写入 Excel
    :param target_tables: 要查找的目标表格名称列表
    :param retries: 重试次数
    """
    if not report_url:
        return None

    for attempt in range(retries):
        try:
            # 打开URL
            driver.get(report_url)

            # 等待页面加载
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.clearit'))
            )

            # 获取页面源码
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
             # 查找包含 “下载公告” 的链接
            download_link = soup.find('a', string='下载公告')
            if download_link and 'href' in download_link.attrs:
                pdf_url = download_link['href']

                # 下载 PDF 文件并保存到 pdfs 文件夹中
                pdf_folder = 'pdfs'
                os.makedirs(pdf_folder, exist_ok=True)  # 确保目录存在
                pdf_path = os.path.join(pdf_folder, f'{sheet_name}.pdf')
                
                pdf_response = requests.get(pdf_url)
                if pdf_response.status_code == 200:
                    with open(pdf_path, 'wb') as pdf_file:
                        pdf_file.write(pdf_response.content)
                    print("PDF 文件下载成功，已保存为:", pdf_path)
                else:
                    print("PDF 文件下载失败，状态码:", pdf_response.status_code)

            # 设置 sheet 名称，随机生成避免冲突
            workbook = writer.book
            worksheet = workbook.create_sheet(title=sheet_name)
            
            current_row = 1

            for target_table_name in target_tables:
                # 寻找目标表格
                target_p_list = soup.find_all('p', string=lambda s: s and target_table_name in s and len(s) < 50)
                
                # 如果未找到匹配的表格名称，去掉“合并”重新查找
                if not target_p_list and "合并" in target_table_name:
                    simplified_name = target_table_name.replace("合并", "")
                    target_p_list = soup.find_all('p', string=lambda s: s and simplified_name in s and len(s) < 50)
                
                if not target_p_list:
                    print(f"未找到匹配的表格名称：{target_table_name}")
                    continue

                target_p = target_p_list[0]  # 取第一个匹配项
                table_title = target_p.get_text(strip=True)
                
                # 将表格标题写入Excel的当前行（不合并单元格）
                cell = worksheet.cell(row=current_row, column=1, value=table_title)
                cell.alignment = Alignment(horizontal='center')
                worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)  # 字体加粗，并设置字体大小为14
                current_row += 1  # 移到下一行准备写入表格数据

                combined_df = pd.DataFrame()

                # 遍历表格
                next_div = target_p.find_next_sibling('div', class_='table-wrap')
                while next_div and 'table-wrap' in next_div.get('class', []):
                    tables = next_div.find_all('table')
                    for table in tables:
                        extracted_html = str(table)
                        html_io = StringIO(extracted_html)
                        df = pd.read_html(html_io)[0]
                        combined_df = pd.concat([combined_df, df], ignore_index=True)
                    next_div = next_div.find_next_sibling()
                    if not (next_div and next_div.name == 'div' and 'table-wrap' in next_div.get('class', [])):
                        break

                # 写入DataFrame内容到Excel
                for row in dataframe_to_rows(combined_df, index=False, header=True):
                    for col, value in enumerate(row, start=1):
                        worksheet.cell(row=current_row, column=col, value=value)
                    current_row += 1  # 每写完一行数据，行号增加

                current_row += 5  # 表格之间的间隔

            # 返回最后生成的标题和内容DataFrame用于检查或进一步使用
            return {'title': sheet_name, 'content': combined_df}

        except ConnectionError as e:
            print(f"连接被拒绝，正在重试 {attempt + 1}/{retries} ... 错误: {e}")
            time.sleep(5)  # 等待 5 秒后重试
        except Exception as e:
            print(f"爬取报告内容时出现错误: {e}")
            break  # 其他错误时停止重试
    
    return None  # 如果重试失败，返回None

# 解析报告中的研发费用和同比数据
def extract_r_d_expenses(report_content):
    if not report_content:
        return None

    # 正则表达式匹配“研发费用”及相关信息
    pattern = re.compile(r"研发费用.*?([\d,.]+).*?同比(增长|下降).*?([\d,.]+)%", re.S)
    match = pattern.search(report_content)

    if match:
        r_d_expense = match.group(1)  # 研发费用
        comparison_type = match.group(2)  # 增长或下降
        comparison_value = match.group(3)  # 同比百分比变化
        return {
            'r_d_expense': r_d_expense,
            'comparison_type': comparison_type,
            'comparison_value': comparison_value
        }
    else:
        return None

# 获取公司股票代码（假设通过公司名称搜索页面）
def get_stock_code_by_company_name(company_name):
    search_url = f"https://so.eastmoney.com/web/s?keyword={company_name}"
    driver.get(search_url)

    try:
        # 等待页面加载完成
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'a.exstock_t_l'))
        )

        # 获取页面内容并解析
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')

        # 根据页面内容提取股票代码
        # 这里需要根据实际页面结构调整选择器
        a_tags = soup.find_all('a', class_='exstock_t_l')
        for a_tag in a_tags:
            text = a_tag.text.strip()
            # 使用正则表达式提取括号内的股票代码
            match = re.search(r'\((\d+)\)', text)
            if match:
                stock_code = match.group(1)  # 提取的股票代码
                print(f"公司: {company_name}, 股票代码: {stock_code}")
                return stock_code
        print(f"公司: {company_name}, 未找到股票代码")
        return None

    except Exception as e:
        print(f"获取公司 {company_name} 股票代码时出现错误: {e}")
        return None

# 示例调用
# companies = ['艾为电子','圣邦股份','恒玄科技','南芯科技','纳芯微','天德钰','中科蓝讯','杰华特','晶丰明源','英集芯','思瑞浦','芯朋微','中微半导','力芯微','必易微','富满微','明微电子','炬芯科技','帝奥微','新相微','希荻微']
companies = ['艾为电子','芯朋微']
# years = [2021, 2022, 2023]
years = [2024,2023]
# target_tables=['合并资产负债表', '合并利润表','合并现金流量表','员工情况']
target_tables=['合并资产负债表', '合并利润表']
# 'zqbg'是中报, 'ndbg'是年报, 'sjdbg'是三季报, 'yjdbg'是一季度报告
# reportTypes = ['zqbg', 'ndbg']
reportTypes = ['sjdbg'];


# 爬取报告并提取研发费用信息
results = crawl_reports_for_companies(companies, years, target_tables, reportTypes)

