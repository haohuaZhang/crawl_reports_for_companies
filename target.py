from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.firefox import GeckoDriverManager
from bs4 import BeautifulSoup
import pandas as pd
import re
import os
import random
import time
from io import StringIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

# 调整所有 sheet 的列宽
def auto_adjust_column_width(excel_file_path):
    workbook = load_workbook(excel_file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 20
            sheet.column_dimensions[column].width = adjusted_width
    workbook.save(excel_file_path)

# 配置 Firefox 浏览器选项
options = Options()
options.set_preference('permissions.default.image', 2)
options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')

# 指定 geckodriver 的路径
service = Service(GeckoDriverManager().install())

# 创建 WebDriver 实例
driver = webdriver.Firefox(service=service, options=options)

# 目标目录
directory = './dist'

# 初始化 reports 文件名的列表
reports_files = []
pattern = r'^targets(\d*)\.xlsx$'

# 查找最大编号的文件
for filename in os.listdir(directory):
    match = re.match(pattern, filename)
    if match:
        num_str = match.group(1)
        number = int(num_str) if num_str else 0
        reports_files.append(number)

max_number = max(reports_files) if reports_files else 0
new_filename = f'targets{max_number + 1}.xlsx'
output_file = os.path.join(directory, new_filename)

# 主函数，爬取多个公司的报告并提取研发费用信息
def crawl_reports_for_companies(companies, years, reportTypes=['zqbg', 'ndbg']):
    results = []
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for company in companies:
            stock_id = get_stock_code_by_company_name(company)
            if not stock_id:
                print(f"无法找到公司 {company} 的股票代码")
                continue

            report_url = f'https://emweb.securities.eastmoney.com/pc_hsf10/pages/index.html?type=web&code={stock_id}&color=b#/cwfx'
            print(f"stock_id {stock_id}  report_url {report_url}")

            report_type = "财务分析报告"
            print(f"正在爬取 {company} 的 {report_type}...")
            sheet_name = f"{company}-{report_type}"
            
            retries = 3  # 设置重试次数
            report = get_report_content_selenium(sheet_name, report_url, writer, years, retries)
            
            if report:
                print(f"{company} 的 {report_type} 报告爬取成功")
                results.append(report)
            else:
                print(f"无法爬取 {company} 的 {report_type} 报告")
            
            time.sleep(random.uniform(3, 7))

    # 如果 Excel 中没有任何 Sheet，则创建一个默认 Sheet
    workbook = writer.book
    if not workbook.sheetnames:
        workbook.create_sheet(title="Default")
    
    # 调整列宽
    auto_adjust_column_width(output_file)
    driver.quit()
    return results

# 使用 Selenium 爬取报告页面内容，带有重试机制
def get_report_content_selenium(sheet_name, report_url, writer, years, retries=5, retry_delay=10):
    if not report_url:
        return None

    for attempt in range(retries):
        try:
            # driver = webdriver.Chrome()  # 启动webdriver
            # 打开URL
            driver.get(report_url)
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'cwfx'))
            )

            cwfx_div = driver.find_element(By.CLASS_NAME, 'cwfx')
            date_tab_li = cwfx_div.find_element(By.XPATH, './/ul[@class="dateTab"]/li[2]')
            date_tab_li.click()
            
            time.sleep(random.uniform(3, 7))
            
            WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.CLASS_NAME, 'zyzb_table'))
            )

            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            tables = soup.select('div.zyzb_table table')
            
            for table in tables:
                if table.get('style') != 'display: none;':
                    rows = table.find_all('tr')
                    
                    # 将数据和标题行标记存储为列表
                    data_rows = []
                    for row in rows:
                        is_title_row = 'title' in row.get('class', [])
                        row_data = [td.get_text(strip=True) for td in row.find_all(['td', 'th'])]
                        data_rows.append((row_data, is_title_row))

                    # 将数据转换为DataFrame
                    df = pd.DataFrame([row_data for row_data, _ in data_rows])
                    df.columns = df.iloc[0]  # 使用第一行作为列名
                    df = df[1:]  # 去除标题行（第一行）

                    # 筛选需要的列
                    columns_to_include = [df.columns[0]]  # 无条件包含第一列
                    for col in df.columns[1:]:
                        col_year_part = str(col).split('-')[0]
                        if any(str(year)[-2:] == col_year_part for year in years):
                            columns_to_include.append(col)

                    filtered_df = df[columns_to_include]  # 过滤列
                    workbook = writer.book
                    
                    try:
                        worksheet = workbook.create_sheet(title=sheet_name)
                    except Exception as e:
                        print(f"创建Sheet失败: {e}")
                        return False
                    
                    # 写入表格名称
                    current_row = 1
                    worksheet.cell(row=current_row, column=1, value=sheet_name).alignment = Alignment(horizontal='center')
                    worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=16)
                    current_row += 1

                    # 写入筛选后的数据，并标红标题行
                    for row_index, (row_data, is_title_row) in enumerate(data_rows):
                        if row_index == 0:
                            # 标题行（列名）
                            for col, value in enumerate(filtered_df.columns, start=1):
                                cell = worksheet.cell(row=current_row, column=col, value=value)
                                if is_title_row:
                                    cell.font = Font(color="FF0000", bold=True)
                            current_row += 1
                        else:
                            # 数据行
                            filtered_row = [row_data[col_index] for col_index, col_name in enumerate(df.columns) if col_name in columns_to_include]
                            for col, value in enumerate(filtered_row, start=1):
                                cell = worksheet.cell(row=current_row, column=col, value=value)
                                if is_title_row:
                                    cell.font = Font(color="FF0000")  # 标红字体
                            current_row += 1
            # driver.quit()
            return True
        except Exception as e:
            print(f"出现错误: {e}, 正在重试 ({attempt + 1}/{retries})")
            time.sleep(retry_delay)
        # finally:
            # driver.quit()
    return False

# 获取公司股票代码
def get_stock_code_by_company_name(company_name):
    search_url = f"https://xueqiu.com/k?q={company_name}"
    driver.get(search_url)
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'p.search__stock__bd__code'))
        )

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        p_tags = soup.find_all('p', class_='search__stock__bd__code')
        for p_tag in p_tags:
            text = p_tag.text.strip()
            if text:
                stock_code = text
                print(f"公司: {company_name}, 股票代码: {stock_code}")
                return stock_code
        print(f"公司: {company_name}, 未找到股票代码")
        return None
    except Exception as e:
        print(f"获取公司 {company_name} 股票代码时出现错误: {e}")
        return None

# 示例调用
companies = ['艾为电子','圣邦股份','恒玄科技','南芯科技','纳芯微','天德钰','中科蓝讯','杰华特','晶丰明源','英集芯','思瑞浦','芯朋微','中微半导','力芯微','必易微','富满微','明微电子','炬芯科技','帝奥微','新相微','希荻微']
# companies = ['艾为电子']
years = [2023, 2022, 2021]
reportTypes = ['ndbg']

results = crawl_reports_for_companies(companies, years, reportTypes)
